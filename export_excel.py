import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import RadarChart, BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule

def salvar_excel(df_ativos, df_valuation=None, df_checklist=None, df_rankings=None, filename="ativos.xlsx"):
    # Cria o arquivo Excel com múltiplas abas
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df_ativos.to_excel(writer, sheet_name="Ativos", index=False)

        if df_valuation is not None:
            df_valuation.to_excel(writer, sheet_name="Valuation", index=False)

        if df_checklist is not None:
            df_checklist.to_excel(writer, sheet_name="Checklist", index=False)

        if df_rankings is not None:
            df_rankings.to_excel(writer, sheet_name="Rankings", index=False)

    # Reabre o arquivo para inserir gráficos e formatações
    wb = load_workbook(filename)

    # === Heatmap Valuation (inclui Lynch_PEG) ===
    if df_valuation is not None:
        ws = wb["Valuation"]
        max_row = len(df_valuation) + 1
        # Supondo colunas: Ticker | Graham | Bazin | Lynch_PEG | AGF | Cotacao
        colunas_heatmap = ["B", "C", "D", "E"]  # Graham, Bazin, Lynch_PEG, AGF
        for col in colunas_heatmap:
            ws.conditional_formatting.add(
                f"{col}2:{col}{max_row}",
                ColorScaleRule(
                    start_type="num", start_value=0, start_color="63BE7B",   # verde
                    mid_type="percentile", mid_value=50, mid_color="FFEB84", # amarelo
                    end_type="num", end_value=1, end_color="F8696B"          # vermelho
                )
            )

    # === Radar comparativo (Top 30) ===
    if df_rankings is not None:
        ws = wb.create_sheet("Radar_Comparativo")
        radar_cols = ["Ticker", "ROE", "Margem_Liquida", "Divida_PL", "Receita_CAGR"]
        df_radar = df_rankings[radar_cols].head(30)
        for r in dataframe_to_rows(df_radar, index=False, header=True):
            ws.append(r)

        chart = RadarChart()
        chart.title = "Radar Comparativo - Top 30"
        chart.style = 26
        chart.y_axis.max = 1
        chart.y_axis.min = 0

        data = Reference(ws, min_col=2, min_row=1, max_col=len(radar_cols), max_row=len(df_radar)+1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(df_radar)+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "H2")

    # === Radar individual (Top 30) ===
    if df_rankings is not None:
        top30 = df_rankings.head(30)
        for _, row in top30.iterrows():
            ticker = row["Ticker"]
            ws = wb.create_sheet(f"Radar_{ticker}")

            indicadores = {
                "ROE": row.get("ROE", 0),
                "Margem_Liquida": row.get("Margem_Liquida", 0),
                "Divida_PL": row.get("Divida_PL", 0),
                "Receita_CAGR": row.get("Receita_CAGR", 0)
            }

            ws.append(["Indicador", "Valor"])
            for k, v in indicadores.items():
                ws.append([k, v])

            chart = RadarChart()
            chart.title = f"Radar de {ticker}"
            chart.style = 26
            chart.y_axis.max = 1
            chart.y_axis.min = 0

            data = Reference(ws, min_col=2, min_row=2, max_row=5)
            cats = Reference(ws, min_col=1, min_row=2, max_row=5)
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(cats)
            ws.add_chart(chart, "D2")

    # === Ranking Setorial Top 10 ===
    if df_rankings is not None and "Setor" in df_ativos.columns:
        ws = wb.create_sheet("Rankings Setoriais")
        ws.append(["Setor", "Ticker", "Score_BH"])

        setores = df_rankings["Setor"].unique()
        for setor in setores:
            top10 = df_rankings[df_rankings["Setor"] == setor].nlargest(10, "Score_BH")
            for _, row in top10.iterrows():
                ws.append([setor, row["Ticker"], row["Score_BH"]])

        # Gráfico de barras por setor (Top 10)
        chart = BarChart()
        chart.title = "Top 10 por Setor - Score Buy & Hold"
        chart.y_axis.title = "Score"
        chart.x_axis.title = "Ticker"

        data = Reference(ws, min_col=3, min_row=2, max_row=ws.max_row)
        cats = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        ws.add_chart(chart, "E2")

    wb.save(filename)
    print(f"✅ Arquivo {filename} gerado com abas, gráficos e Lynch_PEG incluído no Heatmap!")
